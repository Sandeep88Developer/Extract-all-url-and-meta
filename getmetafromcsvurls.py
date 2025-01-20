import requests
from bs4 import BeautifulSoup
import openpyxl

# Load the Excel file containing URLs
input_workbook = openpyxl.load_workbook("scraped_links.xlsx")
input_sheet = input_workbook.active

# Create a new Excel workbook for saving meta data
output_workbook = openpyxl.Workbook()
output_sheet = output_workbook.active
output_sheet.title = "Meta Data"

# Add headers to the output Excel file
output_sheet.cell(row=1, column=1).value = "URL"
output_sheet.cell(row=1, column=2).value = "Meta Title"
output_sheet.cell(row=1, column=3).value = "Meta Description"

# Start reading from the second row of the input file (to skip headers)
row = 2
output_row = 2

for cell in input_sheet['A']:
    if cell.row == 1:  # Skip the header row
        continue

    url = cell.value
    if url:
        try:
            # Fetch the page and parse the HTML
            response = requests.get(url, timeout=10)
            soup = BeautifulSoup(response.text, 'html.parser')

            # Extract meta title and description
            meta_title = soup.title.string.strip() if soup.title else "No Title"
            meta_description = soup.find("meta", attrs={"name": "description"})
            meta_description = meta_description['content'].strip() if meta_description else "No Description"

            # Write data to the output Excel file
            output_sheet.cell(row=output_row, column=1).value = url
            output_sheet.cell(row=output_row, column=2).value = meta_title
            output_sheet.cell(row=output_row, column=3).value = meta_description
            output_row += 1

        except Exception as e:
            print(f"Error processing URL {url}: {e}")
            # Log the error in the output file for the problematic URL
            output_sheet.cell(row=output_row, column=1).value = url
            output_sheet.cell(row=output_row, column=2).value = "Error"
            output_sheet.cell(row=output_row, column=3).value = str(e)
            output_row += 1

# Save the output Excel file
output_workbook.save("meta_data_output.xlsx")
print("Meta data saved to 'meta_data_output.xlsx'")
