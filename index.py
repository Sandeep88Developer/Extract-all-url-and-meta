import requests
from bs4 import BeautifulSoup
import openpyxl

# Initialize an Excel workbook and sheet
workbook = openpyxl.Workbook()
sheet = workbook.active
sheet.title = "Scraped Links"

# Add headers to the Excel file
sheet.cell(row=1, column=1).value = "Links"

# URL to scrape
urls = 'https://xyz.com//'  #website url 
grab = requests.get(urls)
soup = BeautifulSoup(grab.text, 'html.parser')

# Start inserting data from row 2
row = 2

# Traverse paragraphs and write to Excel
for link in soup.find_all("a"):
    data = link.get('href')
    if data:  # Ensure the href is not None
        sheet.cell(row=row, column=1).value = data
        row += 1

# Save the Excel file
workbook.save("scraped_links.xlsx")
print("Data saved to 'scraped_links.xlsx'")
